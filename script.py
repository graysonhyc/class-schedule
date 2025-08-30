import io
import re
import csv
from datetime import time
from typing import Dict, List, Tuple, Set

import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

# ----------------- utils -----------------

DAY_MAP = {"一": "星期一", "二": "星期二", "三": "星期三", "四": "星期四", "五": "星期五"}

def _clean(x) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    if s in {"-", "—", "–", "－"} or s.strip(" -—–－") == "":
        return ""
    return re.sub(r"\s+", " ", s)

def parse_start_time(label: str):
    """Parse '9:05am-9:40am' -> time(9,5)"""
    if not isinstance(label, str):
        return None
    m = re.match(r"\s*(\d{1,2}):(\d{2})(am|pm)?\s*-\s*(\d{1,2}):(\d{2})(am|pm)?", label, re.I)
    if not m:
        return None
    h = int(m.group(1)); mi = int(m.group(2)); ap = (m.group(3) or "").lower()
    if ap == "pm" and h < 12: h += 12
    if ap == "am" and h == 12: h = 0
    return time(h, mi)

def find_class_columns(ws: Worksheet) -> Dict[str, int]:
    """Row 1 has headers like '1A','1B'... Data is in the SAME (left) column of each merged pair."""
    class_cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if isinstance(v, str) and re.fullmatch(r"\d+[A-E]", v):
            class_cols[v] = c
    return class_cols

def build_mapping(file_a_bytes: bytes) -> Dict[str, Dict[time, Dict[str, str]]]:
    """
    From school_timetable.xlsx build:
      map[day_name][start_time][class] = "(科目 老師)"
    Col A = period # rows (1..10...), Col B = start time,
    subject on the numbered row, teacher at +2 rows.
    """
    wb = load_workbook(io.BytesIO(file_a_bytes), data_only=True)
    result: Dict[str, Dict[time, Dict[str, str]]] = {}
    for ws in wb.worksheets:
        class_cols = find_class_columns(ws)
        day_map: Dict[time, Dict[str, str]] = {}
        r = 1
        while r <= ws.max_row:
            a = ws.cell(r, 1).value
            if isinstance(a, int):
                start_t = ws.cell(r, 2).value  # datetime.time
                if start_t:
                    for cls, col in class_cols.items():
                        subject = _clean(ws.cell(r, col).value)
                        teacher = _clean(ws.cell(r + 2, col).value) if r + 2 <= ws.max_row else ""
                        if not subject and not teacher:
                            continue
                        suffix = f"({subject} {teacher})".strip()
                        suffix = re.sub(r"\s{2,}", " ", suffix)
                        day_map.setdefault(start_t, {})[cls] = suffix
                r += 1
            else:
                r += 1
        result[ws.title] = day_map
    return result

CLASS_TOKEN = re.compile(r"(\d+[A-E])(?:\d+)?")  # class (1A..6E) with optional student number

def annotate_schedule(wb: Workbook, mapping: Dict[str, Dict[time, Dict[str, str]]]) -> Tuple[int, List[Tuple]]:
    """
    Append '(科目 老師)' to each line that contains a class token, based on (day, start_time, class).
    Returns (cells_updated, unmatched_rows).
    """
    ws = wb.active
    # row 2 has '一 二 三 四 五'
    day_headers = [ws.cell(2, c).value for c in range(2, ws.max_column + 1)]
    day_names = [DAY_MAP.get(v, v) for v in day_headers]

    unmatched: List[Tuple] = []
    changed = 0

    for r in range(3, ws.max_row + 1):
        start_t = parse_start_time(ws.cell(r, 1).value)
        if not start_t:
            continue
        for c_idx, day_name in enumerate(day_names, start=2):
            cell = ws.cell(r, c_idx)
            text = cell.value
            if not isinstance(text, str) or not text.strip():
                continue

            lines = text.splitlines()
            new_lines = []
            for line in lines:
                m = CLASS_TOKEN.search(line)
                if not m:
                    new_lines.append(line); continue
                cls = m.group(1)
                suffix = mapping.get(day_name, {}).get(start_t, {}).get(cls)
                if not suffix:
                    unmatched.append((ws.title, r, cell.coordinate, day_name, start_t.isoformat(), cls, line))
                    new_lines.append(line)
                else:
                    # append on the same line (matches your DOCX)
                    if suffix in line:
                        new_lines.append(line)
                    else:
                        new_lines.append(f"{line} {suffix}")

            new_text = "\n".join(new_lines)
            if new_text != text:
                cell.value = new_text
                changed += 1

    return changed, unmatched

# ----------------- UI -----------------

st.set_page_config(page_title="📚 ST Schedule Annotator", layout="centered")
st.title("📚 言語治療時間表自動標註")
st.caption("上傳「每週課表」與「九月日程」Excel，系統會根據星期＋開始時間＋班別自動加上 (科目 老師) 後綴。")

file_a = st.file_uploader("上傳 school_timetable.xlsx（含 星期一~五 5 個工作表）", type=["xlsx"], key="a")
file_b = st.file_uploader("上傳 september_st_timetable.xlsx（待標註）", type=["xlsx"], key="b")

if st.button("開始標註", type="primary", disabled=not (file_a and file_b)):
    try:
        with st.spinner("讀取並建立對照表…"):
            timetable_map = build_mapping(file_a.read())
            total_keys = sum(len(v) for v in timetable_map.values())
            st.success(f"完成：{total_keys} 個時間段載入。")

        file_b.seek(0)
        wb = load_workbook(file_b, data_only=True)

        with st.spinner("套用標註…"):
            changed, unmatched = annotate_schedule(wb, timetable_map)

        # Export outputs
        out_xlsx = io.BytesIO()
        wb.save(out_xlsx); out_xlsx.seek(0)

        csv_buf = io.StringIO()
        w = csv.writer(csv_buf); w.writerow(["sheet","row","cell","day","start_time","class","line"])
        for row in unmatched: w.writerow(row)
        csv_bytes = csv_buf.getvalue().encode("utf-8")

        st.success(f"已更新 {changed} 個儲存格。未匹配：{len(unmatched)}")
        st.download_button("⬇️ 下載已標註 Excel", data=out_xlsx,
                           file_name="september_st_timetable_annotated.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.download_button("⬇️ 下載 unmatched_keys.csv", data=csv_bytes,
                           file_name="unmatched_keys.csv", mime="text/csv")

        if st.checkbox("查看部分對照表（預覽 30 條）"):
            sample = {}
            for day, by_time in timetable_map.items():
                for t, by_cls in by_time.items():
                    for cls, suf in by_cls.items():
                        sample[f"{day} {t} {cls}"] = suf
                        if len(sample) >= 30: break
                    if len(sample) >= 30: break
                if len(sample) >= 30: break
            st.write(sample)

    except Exception as e:
        st.error("處理失敗")
        st.exception(e)

st.markdown("---")
st.markdown("""
**要點修正**
- 以「星期 + 開始時間 + 班別」對應科目與老師（非班別後的數字，因那是學生編號）。
- 每個節次為 3 行：第 1 行科目、第 3 行老師（中間是 `-`）；從 Column B 讀開始時間。
- 標註時逐行處理同一儲存格的多位學生，為每一行加上各自的後綴。
- 班別標頭使用合併儲存格的**左列**。
""")